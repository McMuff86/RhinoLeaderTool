#! python 3
# -*- coding: utf-8 -*-
import rhinoscriptsyntax as rs
import Rhino
import Rhino.FileIO as FileIO
import Rhino.UI
import scriptcontext as sc
import os
import csv
import json
from datetime import datetime

def read_csv_attributes(path):
    attributes = {}
    try:
        with open(path, mode="r", encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if len(row) == 2:
                    key, value = row
                    attributes[key.strip()] = value.strip()
    except Exception as e:
        print("Fehler beim Lesen der CSV:", e)
    return attributes

def get_base_path():
    user_dir = os.path.expanduser("~")
    return os.path.join(user_dir, "source", "repos", "work", "library", "RhinoLeaderTool")

def load_config():
    base_path = get_base_path()
    config_path = os.path.join(base_path, "config.json")
    default_config = {
        "base_path": base_path,
        "template_path": "LeaderAnnotationTemplate.3dm",
        "logging": {
            "mode": "csv",  # csv | xlsx
            "file": "created_leaders_log.csv"
        },
        "types": {
            "rahmentuere": {"dimstyle": "Standard 1:10 Rahmenbeschriftung", "csv": "rahmentuere.csv"},
            "zargentuere": {"dimstyle": "Standard 1:10 Zargenbeschriftung", "csv": "zargentuere.csv"},
            "schiebetuere": {"dimstyle": "Standard 1:10 Schiebetürbeschriftung", "csv": "schiebetuere.csv"},
            "rahmentuere_w": {"dimstyle": "Standard 1:10 Rahmenbeschriftung WHG Eingang", "csv": "rahmentuerew.csv"},
            "spez": {"dimstyle": "Standard 1:10 Spez.Rahmenbeschriftung", "csv": "spez.csv"}
        }
    }

    try:
        if os.path.isfile(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                file_cfg = json.load(f)
            # flach mergen (Datei-Werte überschreiben Default)
            for k, v in file_cfg.items():
                default_config[k] = v
    except Exception as e:
        print("Konnte config.json nicht laden (verwende Defaults):", e)
    return default_config

def apply_usertext_overrides(cfg, data):
    try:
        overrides = (cfg.get("defaults", {}) or {}).get("usertext_overrides", {}) or {}
        na_value = (cfg.get("export", {}) or {}).get("na_value", "NA")
        for k, v in overrides.items():
            current = data.get(k)
            if current is None or str(current).strip() == "" or str(current).strip().upper() == str(na_value).upper():
                data[k] = v
    except Exception:
        pass

def get_na_value(cfg):
    try:
        return (cfg.get("export", {}) or {}).get("na_value", "NA")
    except Exception:
        return "NA"

def load_saved_globals(cfg):
    saved = {}
    try:
        section = "RhinoLeaderToolGlobals"
        prompt_keys = (cfg.get("defaults", {}) or {}).get("prompt_keys", [])
        for k in prompt_keys:
            try:
                v = rs.GetDocumentData(section, k)
                if v is not None:
                    saved[k] = v
            except Exception:
                pass
    except Exception:
        pass
    return saved

def save_globals(cfg, data):
    try:
        section = "RhinoLeaderToolGlobals"
        for k, v in (data or {}).items():
            try:
                rs.SetDocumentData(section, k, str(v) if v is not None else "")
            except Exception:
                pass
    except Exception:
        pass

def prefill_data_with_saved_globals(cfg, data):
    try:
        saved = load_saved_globals(cfg)
        for k, v in saved.items():
            data[k] = v
    except Exception:
        pass

def merge_globals_into_data(cfg, data, globals_dict):
    try:
        na = get_na_value(cfg)
        for k, v in (globals_dict or {}).items():
            cur = data.get(k)
            if cur is None or str(cur).strip() == "" or str(cur).strip().upper() == str(na).upper():
                data[k] = v
    except Exception:
        pass

def maybe_prompt_for_globals(cfg, data):
    try:
        defaults_cfg = cfg.get("defaults", {}) or {}
        always_prompt = bool(defaults_cfg.get("always_prompt", False))
        prompt_on_first = bool(defaults_cfg.get("prompt_on_first_leader", False))
        if not (always_prompt or prompt_on_first):
            return
        # Prüfen, ob bereits irgendein Leader im Dokument existiert
        any_leader = False
        for obj in sc.doc.Objects:
            try:
                if isinstance(obj.Geometry, Rhino.Geometry.Leader):
                    any_leader = True
                    break
            except Exception:
                continue
        if (any_leader and not always_prompt) and prompt_on_first:
            return
        # Prompt für definierte Keys
        prompt_keys = list(defaults_cfg.get("prompt_keys") or [])
        na_value = get_na_value(cfg)
        # Vorbelegen aus gespeicherten Dokumentwerten
        prefill_data_with_saved_globals(cfg, data)
        # Bevorzugt: Eto-Dialog (Rhino 8, CPython + pythonnet)
        try:
            import Eto.Forms as forms
            import Eto.Drawing as drawing

            dialog = forms.Dialog()
            dialog.Title = "Globale Werte setzen"

            layout = forms.DynamicLayout()
            layout.Spacing = drawing.Size(6, 6)
            layout.Padding = drawing.Padding(10)

            textboxes = {}
            for key in prompt_keys:
                lbl = forms.Label()
                lbl.Text = key
                tb = forms.TextBox()
                tb.Text = str(data.get(key, na_value))
                textboxes[key] = tb
                layout.AddRow(lbl, tb)

            ok_btn = forms.Button()
            ok_btn.Text = "OK"
            cancel_btn = forms.Button()
            cancel_btn.Text = "Abbrechen"
            layout.AddSeparateRow(None, ok_btn, cancel_btn)

            def on_ok(sender, e):
                dialog.Tag = True
                dialog.Close()

            def on_cancel(sender, e):
                dialog.Tag = False
                dialog.Close()

            ok_btn.Click += on_ok
            cancel_btn.Click += on_cancel

            dialog.Content = layout
            dialog.Tag = False
            try:
                Rhino.UI.EtoExtensions.ShowSemiModal(dialog, sc.doc, Rhino.UI.RhinoEtoApp.MainWindow)
            except Exception as semimodal_ex:
                print("Hinweis: EtoExtensions.ShowSemiModal fehlgeschlagen:", semimodal_ex)
                dialog.ShowModal(Rhino.UI.RhinoEtoApp.MainWindow)

            if dialog.Tag:
                for key, tb in textboxes.items():
                    val = (tb.Text or "").strip()
                    if val != "":
                        data[key] = val
            # Nach Bestätigung speichern
            save_globals(cfg, {k: data.get(k, na_value) for k in prompt_keys})
            return
        except Exception as eto_ex:
            print("Eto-Dialog nicht verfügbar, verwende GetString-Fallback. Grund:", eto_ex)
            # Fallback: einfache Texteingaben
            for k in prompt_keys:
                existing = data.get(k, na_value)
                try:
                    val = rs.GetString(f"Set value for {k} (Enter=keep)", str(existing))
                    if val is not None and val != "":
                        data[k] = val
                except Exception as gs_ex:
                    print("GetString fehlgeschlagen für", k, "Grund:", gs_ex)
            save_globals(cfg, {k: data.get(k, na_value) for k in prompt_keys})
    except Exception:
        pass

def maybe_prompt_for_type_specific(cfg, typ, data):
    try:
        defaults_cfg = cfg.get("defaults", {}) or {}
        keys = list(defaults_cfg.get("type_specific_keys") or [])
        if not keys:
            return
        na_value = get_na_value(cfg)
        # pro Typ einmalig fragen? Wir nutzen DocData mit Typ-Präfix
        section = f"RhinoLeaderToolType:{typ}"
        missing = []
        for k in keys:
            cur = data.get(k, na_value)
            if cur is None or str(cur).strip() == "" or str(cur).strip().upper() == str(na_value).upper():
                # schauen ob bereits im Dokument pro Typ gespeichert
                saved = rs.GetDocumentData(section, k)
                if saved is not None and saved.strip() != "":
                    data[k] = saved
                else:
                    missing.append(k)
        if not missing:
            return
        # Eto-Dialog für die fehlenden type-spezifischen Keys
        try:
            import Eto.Forms as forms
            import Eto.Drawing as drawing

            dialog = forms.Dialog()
            dialog.Title = f"{typ}: Werte setzen"
            layout = forms.DynamicLayout()
            layout.Spacing = drawing.Size(6, 6)
            layout.Padding = drawing.Padding(10)
            textboxes = {}
            for key in missing:
                lbl = forms.Label()
                lbl.Text = key
                tb = forms.TextBox()
                tb.Text = str(data.get(key, na_value))
                textboxes[key] = tb
                layout.AddRow(lbl, tb)
            ok_btn = forms.Button(); ok_btn.Text = "OK"
            cancel_btn = forms.Button(); cancel_btn.Text = "Abbrechen"
            layout.AddSeparateRow(None, ok_btn, cancel_btn)
            def on_ok(sender, e):
                dialog.Tag = True; dialog.Close()
            def on_cancel(sender, e):
                dialog.Tag = False; dialog.Close()
            ok_btn.Click += on_ok; cancel_btn.Click += on_cancel
            dialog.Content = layout; dialog.Tag = False
            try:
                Rhino.UI.EtoExtensions.ShowSemiModal(dialog, sc.doc, Rhino.UI.RhinoEtoApp.MainWindow)
            except Exception as semimodal_ex:
                print("Hinweis: SemiModal fehlgeschlagen:", semimodal_ex)
                dialog.ShowModal(Rhino.UI.RhinoEtoApp.MainWindow)
            if dialog.Tag:
                for key, tb in textboxes.items():
                    val = (tb.Text or "").strip()
                    if val != "":
                        data[key] = val
                        rs.SetDocumentData(section, key, val)
            return
        except Exception as eto_ex:
            print("Eto-Dialog nicht verfügbar (type-specific), Fallback:", eto_ex)
            for k in missing:
                existing = data.get(k, na_value)
                try:
                    val = rs.GetString(f"{typ} – {k}", str(existing))
                    if val is not None and val != "":
                        data[k] = val
                        rs.SetDocumentData(section, k, val)
                except Exception as gs_ex:
                    print("GetString fehlgeschlagen für", k, "Grund:", gs_ex)
    except Exception:
        pass

def get_dimstyle_id(name):
    dimstyle = sc.doc.DimStyles.FindName(name)
    if dimstyle:
        return dimstyle.Id
    else:
        print(f"Bemaßungsstil '{name}' nicht gefunden.")
        return None

def import_dimstyles_from_template(template_path):
    try:
        if os.path.isfile(template_path):
            model = FileIO.File3dm.Read(template_path)
            if model is None:
                raise Exception("Template konnte nicht gelesen werden.")
            imported_count = 0
            for ds in model.DimStyles:
                if sc.doc.DimStyles.FindName(ds.Name) is None:
                    # Kopie anlegen und hinzufügen
                    dup = ds.Duplicate()
                    index = sc.doc.DimStyles.Add(dup)
                    if index is not None and index >= 0:
                        imported_count += 1
            if imported_count > 0:
                print(f"{imported_count} DimStyles aus Template importiert.")
            return True
        else:
            print("Template-Datei nicht gefunden:", template_path)
    except Exception as e:
        print("Fehler beim Import der DimStyles aus Template:", e)
    return False

def ensure_dimstyle_exists(dimstyle_name, template_full_path):
    dim_id = get_dimstyle_id(dimstyle_name)
    if dim_id:
        return dim_id
    # Versuch, aus Template zu importieren
    if import_dimstyles_from_template(template_full_path):
        dim_id = get_dimstyle_id(dimstyle_name)
        if dim_id:
            return dim_id
    return None

def create_leader_with_style(dimstyle_id):
    rs.Command("_Leader _Pause _Pause _Pause _Enter", echo=False)
    objs = rs.LastCreatedObjects()
    if not objs:
        print("Kein Leader erstellt.")
        return None

    for obj in objs:
        if rs.ObjectType(obj) == 512:
            rhobj = sc.doc.Objects.Find(obj)
            geo = rhobj.Geometry
            geo.DimensionStyleId = dimstyle_id
            sc.doc.Objects.Replace(obj, geo)
            return obj
    return None

def attach_usertext(obj_id, data):
    for key, value in data.items():
        rs.SetUserText(obj_id, key, value)
    print(f"{len(data)} UserText-Einträge hinzugefügt.")
    try:
        # Persistente Zuordnung: LeaderGUID als UserText speichern
        existing = rs.GetUserText(obj_id, "LeaderGUID")
        if existing is None or existing == "":
            rs.SetUserText(obj_id, "LeaderGUID", str(obj_id))
        # Schema-Version setzen, falls nicht vorhanden
        schema = rs.GetUserText(obj_id, "SchemaVersion")
        if schema is None or schema == "":
            rs.SetUserText(obj_id, "SchemaVersion", "1.0")
    except Exception:
        pass

def log_leader_creation(cfg, leader_id, typ, dimstyle_name, csv_filename):
    try:
        log_cfg = cfg.get("logging", {})
        mode = (log_cfg.get("mode") or "csv").lower()
        cfg_base = cfg.get("base_path")
        base_path = cfg_base if (cfg_base and os.path.isdir(cfg_base)) else get_base_path()
        log_file = log_cfg.get("file") or ("created_leaders_log.csv" if mode == "csv" else "created_leaders_log.xlsx")
        # Korrigiere Dateiendung, falls XLSX-Modus mit unpassender Endung
        if mode == "xlsx" and not log_file.lower().endswith((".xlsx", ".xlsm")):
            root, _ = os.path.splitext(log_file)
            log_file = root + ".xlsx"
        log_path = os.path.join(base_path, log_file)

        timestamp = datetime.now().isoformat(timespec="seconds")
        doc_name = sc.doc.Name or "Untitled"
        leader_str = str(leader_id)

        if mode == "xlsx":
            try:
                from openpyxl import Workbook, load_workbook
                if os.path.exists(log_path):
                    wb = load_workbook(log_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["timestamp", "document", "leader_guid", "type", "dimstyle", "csv"])
                ws.append([timestamp, doc_name, leader_str, typ, dimstyle_name, csv_filename])
                wb.save(log_path)
                return
            except Exception as e:
                print("Excel-Logging nicht verfügbar, fallback zu CSV:", e)
                mode = "csv"

        # CSV (Default/Fallback)
        import io
        header_needed = not os.path.exists(log_path)
        with io.open(log_path, "a", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            if header_needed:
                writer.writerow(["timestamp", "document", "leader_guid", "type", "dimstyle", "csv"])
            writer.writerow([timestamp, doc_name, leader_str, typ, dimstyle_name, csv_filename])
    except Exception as e:
        print("Fehler beim Logging der Leader-Erstellung:", e)

def select_preset(cfg, typ):
    entry = (cfg.get("types", {}) or {}).get(typ, {})
    presets = entry.get("presets") or []
    default_preset = entry.get("default_preset")
    if not presets:
        # Kein Preset definiert → CSV aus Typ und Preset-Label auf "Standard" oder default_preset
        preset_label = default_preset or "Standard"
        return entry.get("csv"), preset_label
    # Prompt zur Auswahl eines Presets (namenbasiert)
    names = [p.get("name") for p in presets if p.get("name")]
    if not names:
        return entry.get("csv"), (default_preset or "Standard")
    default_name = default_preset if (default_preset in names) else names[0]
    try:
        selection = rs.ListBox(names, message=f"Select preset for {typ}", title="Choose Preset", default=default_name)
        chosen_name = selection or default_name
        for p in presets:
            if p.get("name") == chosen_name:
                return p.get("csv") or entry.get("csv"), chosen_name
        return entry.get("csv"), (chosen_name or default_name or "Standard")
    except Exception:
        return entry.get("csv"), (default_name or default_preset or "Standard")


def run_leader_for_type(typ):
    cfg = load_config()
    types_cfg = cfg.get("types", {})
    if typ not in types_cfg:
        print(f"Typ '{typ}' nicht erkannt.")
        return

    dimstyle_name = types_cfg[typ]["dimstyle"]
    # Zuerst globale Werte (Dialog), danach Preset auswählen
    globals_buf = {}
    maybe_prompt_for_globals(cfg, globals_buf)
    # CSV-Datei ggf. via Preset-Auswahl überschreiben (und Preset-Name zurückgeben)
    csv_filename, preset_name = select_preset(cfg, typ)

    cfg_base = cfg.get("base_path")
    base_path = cfg_base if (cfg_base and os.path.isdir(cfg_base)) else get_base_path()
    csv_path = os.path.join(base_path, csv_filename)
    template_rel = cfg.get("template_path") or "LeaderAnnotationTemplate.3dm"
    template_path = template_rel if os.path.isabs(template_rel) else os.path.join(base_path, template_rel)

    csv_data = read_csv_attributes(csv_path)
    # Preset-Name als UserText mitschreiben (für spätere Filterung) – immer setzen
    try:
        csv_data["Preset"] = preset_name or "Standard"
    except Exception:
        pass
    # LeaderType (interner Typ) mitschreiben, um spätere Synchronisation zu ermöglichen
    try:
        csv_data["LeaderType"] = typ
    except Exception:
        pass
    # Gespeicherte globale Werte in CSV-Daten übernehmen, dann Overrides anwenden
    saved_globals = load_saved_globals(cfg)
    merge_globals_into_data(cfg, csv_data, saved_globals)
    # Typ-spezifische Keys erfragen (z. B. Betriebsauftrag) und mergen
    type_specific = {}
    maybe_prompt_for_type_specific(cfg, typ, type_specific)
    merge_globals_into_data(cfg, csv_data, type_specific)
    apply_usertext_overrides(cfg, csv_data)
    style_id = ensure_dimstyle_exists(dimstyle_name, template_path)
    if style_id:
        leader_id = create_leader_with_style(style_id)
        if leader_id:
            attach_usertext(leader_id, csv_data)
            log_leader_creation(cfg, leader_id, typ, dimstyle_name, csv_filename)
